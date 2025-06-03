import streamlit as st
import pandas as pd
import qrcode
import io
from PIL import Image
import os
import tempfile
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sqlite3
from datetime import datetime
import base64
from supabase import create_client
import json
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Initialize Supabase client
supabase = create_client(
    os.getenv("SUPABASE_URL"),
    os.getenv("SUPABASE_KEY")
)

# Page configuration
st.set_page_config(
    page_title="Système de Gestion des QR Codes",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
    <style>
    .main {
        padding: 2rem;
    }
    .stButton>button {
        width: 100%;
        background-color: #366092;
        color: white;
        border-radius: 5px;
        padding: 0.5rem 1rem;
        border: none;
    }
    .stButton>button:hover {
        background-color: #2d4d7a;
    }
    .search-box {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 5px;
        margin-bottom: 1rem;
    }
    .success-box {
        background-color: #d4edda;
        padding: 1rem;
        border-radius: 5px;
        margin-bottom: 1rem;
    }
    .data-box {
        background-color: #ffffff;
        padding: 1rem;
        border-radius: 5px;
        margin-bottom: 1rem;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .metric-box {
        background-color: #f8f9fa;
        padding: 1rem;
        border-radius: 5px;
        margin: 0.5rem;
        text-align: center;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 2rem;
    }
    .stTabs [data-baseweb="tab"] {
        height: 4rem;
        white-space: pre-wrap;
        background-color: #f0f2f6;
        border-radius: 4px 4px 0 0;
        gap: 1rem;
        padding-top: 10px;
        padding-bottom: 10px;
    }
    .stTabs [aria-selected="true"] {
        background-color: #366092;
        color: white;
    }
    </style>
    """, unsafe_allow_html=True)

# Column mapping
COLUMN_MAPPING = {
    "PH": "ph",
    "DTR": "dtr",
    "nombre de planche": "nombre_planche",
    "numero de planche": "numero_planche",
    "ligne": "ligne",
    "position": "position",
    "niveau": "niveau"
}

def get_all_records():
    try:
        response = supabase.table('emplacements').select("*").execute()
        return pd.DataFrame(response.data)
    except Exception as e:
        st.error(f"Erreur lors de la récupération des données: {str(e)}")
        return pd.DataFrame()

def get_statistics(df):
    if df.empty:
        return {
            "Total Records": 0,
            "Unique PH": 0,
            "Unique DTR": 0,
            "Total Planches": 0
        }
    
    stats = {
        "Total Records": len(df),
        "Unique PH": df['ph'].nunique(),
        "Unique DTR": df['dtr'].nunique(),
        "Total Planches": df['nombre_planche'].sum() if 'nombre_planche' in df.columns else 0
    }
    return stats

def create_qr_code(data, filename, size=200):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4
    )
    qr.add_data(data)
    qr.make(fit=True)
    qr_image = qr.make_image(fill_color="black", back_color="white")
    qr_image = qr_image.resize((size, size), Image.Resampling.LANCZOS)
    temp_dir = tempfile.mkdtemp()
    filepath = os.path.join(temp_dir, filename)
    qr_image.save(filepath, format="PNG", quality=95)
    return filepath

def create_excel_with_qr_codes(data, filename, is_emplacement=True):
    wb = Workbook()
    ws = wb.active
    ws.title = "Emplacements" if is_emplacement else "Planches"
    
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['PH', 'DTR', 'QR Code'] if is_emplacement else ['Ligne', 'Position', 'Niveau', 'QR Code']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data[:-1], 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        qr_path = row_data[-1]
        img = XLImage(qr_path)
        qr_col = len(headers)
        cell = ws.cell(row=row_idx, column=qr_col)
        ws.row_dimensions[row_idx].height = 150
        ws.add_image(img, f"{get_column_letter(qr_col)}{row_idx}")
    
    ws.column_dimensions[get_column_letter(len(headers))].width = 30
    wb.save(filename)
    return filename

def search_records(ph=None, dtr=None):
    try:
        query = supabase.table('emplacements').select("*")
        
        if ph:
            query = query.ilike('ph', f'%{ph}%')
        if dtr:
            query = query.ilike('dtr', f'%{dtr}%')
            
        response = query.execute()
        return pd.DataFrame(response.data)
    except Exception as e:
        st.error(f"Erreur lors de la recherche: {str(e)}")
        return pd.DataFrame()

# Sidebar
with st.sidebar:
    st.image("https://img.icons8.com/color/96/000000/qr-code--v1.png", width=100)
    st.markdown("## Menu")
    st.markdown("---")
    
    # Quick stats in sidebar
    try:
        all_data = get_all_records()
        if not all_data.empty:
            stats = get_statistics(all_data)
            st.markdown("### 📊 Statistiques rapides")
            for key, value in stats.items():
                st.metric(label=key, value=value)
    except:
        pass

# Main title with custom styling
st.markdown("<h1 style='text-align: center; color: #366092;'>Système de Gestion des QR Codes</h1>", unsafe_allow_html=True)

# Create tabs
tab1, tab2, tab3 = st.tabs(["📤 Importation", "🔍 Recherche", "📊 Données"])

with tab1:
    st.markdown("### Importation de données")
    uploaded_file = st.file_uploader("Choisissez le fichier Excel d'entrée", type=['xlsx', 'xls'])
    
    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file)
            required_columns = list(COLUMN_MAPPING.keys())
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                st.error(f"Colonnes manquantes: {', '.join(missing_columns)}")
            else:
                # Rename columns to match database schema
                df = df.rename(columns=COLUMN_MAPPING)
                
                # Save to Supabase
                for _, row in df.iterrows():
                    data = row.to_dict()
                    supabase.table('emplacements').insert(data).execute()
                
                st.success("Données importées avec succès!")
                
                # Generate QR codes and Excel files
                temp_dir = tempfile.mkdtemp()
                emplacement_data = []
                planche_data = []
                
                for index, row in df.iterrows():
                    emplacement_qr_data = f"PH:{row['ph']}\nDTR:{row['dtr']}\nnb_planche:{row['nombre_planche']}\nnum_planche:{row['numero_planche']}\nligne:{row['ligne']}\nposition:{row['position']}\nniveau:{row['niveau']}"
                    emplacement_qr_path = create_qr_code(emplacement_qr_data, f"emplacement_{index}.png")
                    
                    emplacement_data.append([
                        row['ph'],
                        row['dtr'],
                        emplacement_qr_path
                    ])
                    
                    planche_qr_data = f"ligne:{row['ligne']}\nposition:{row['position']}\nniveau:{row['niveau']}"
                    planche_qr_path = create_qr_code(planche_qr_data, f"planche_{index}.png")
                    
                    planche_data.append([
                        row['ligne'],
                        row['position'],
                        row['niveau'],
                        planche_qr_path
                    ])
                
                emplacement_file = create_excel_with_qr_codes(emplacement_data, "feuille_emplacement.xlsx", True)
                planche_file = create_excel_with_qr_codes(planche_data, "feuille_planche.xlsx", False)
                
                col1, col2 = st.columns(2)
                with col1:
                    with open(emplacement_file, 'rb') as f:
                        st.download_button(
                            label="📥 Télécharger fichier Emplacements",
                            data=f,
                            file_name="feuille_emplacement.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                with col2:
                    with open(planche_file, 'rb') as f:
                        st.download_button(
                            label="📥 Télécharger fichier Planches",
                            data=f,
                            file_name="feuille_planche.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                # Cleanup
                for file in os.listdir(temp_dir):
                    os.remove(os.path.join(temp_dir, file))
                os.rmdir(temp_dir)
                os.remove(emplacement_file)
                os.remove(planche_file)
                
        except Exception as e:
            st.error(f"Erreur lors du traitement: {str(e)}")

with tab2:
    st.markdown("### Recherche de données")
    st.markdown('<div class="search-box">', unsafe_allow_html=True)
    
    search_col1, search_col2 = st.columns(2)
    with search_col1:
        ph_search = st.text_input("Rechercher par PH")
    with search_col2:
        dtr_search = st.text_input("Rechercher par DTR")
    
    if st.button("🔎 Rechercher"):
        results = search_records(ph_search, dtr_search)
        if not results.empty:
            st.markdown('<div class="success-box">', unsafe_allow_html=True)
            st.dataframe(results)
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.warning("Aucun résultat trouvé")
    
    st.markdown('</div>', unsafe_allow_html=True)

with tab3:
    st.markdown("### Visualisation des données")
    
    # Statistics section
    st.markdown("#### 📊 Statistiques")
    try:
        all_data = get_all_records()
        if not all_data.empty:
            stats = get_statistics(all_data)
            
            # Display metrics in a grid
            cols = st.columns(4)
            for i, (key, value) in enumerate(stats.items()):
                with cols[i % 4]:
                    st.metric(label=key, value=value)
            
            # Data summary
            st.markdown("#### 📋 Résumé des données")
            st.markdown('<div class="data-box">', unsafe_allow_html=True)
            st.dataframe(all_data.describe())
            st.markdown('</div>', unsafe_allow_html=True)
            
            # Full data view
            st.markdown("#### 📋 Toutes les données")
            st.markdown('<div class="data-box">', unsafe_allow_html=True)
            st.dataframe(all_data)
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.info("Aucune donnée disponible dans la base de données")
    except Exception as e:
        st.warning("Erreur lors de la récupération des données")

# Footer
st.markdown("---")
st.markdown("<p style='text-align: center; color: #666;'>© 2024 Système de Gestion des QR Codes</p>", unsafe_allow_html=True) 